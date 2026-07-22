"""Build compact Fall 2024 IPEDS first-time undergraduate residence measures."""

from __future__ import annotations

import csv
import io
import json
import tempfile
import urllib.request
import zipfile
from collections import defaultdict
from io import BytesIO
from pathlib import Path


PROJECT = Path(__file__).resolve().parents[1]
DATA = PROJECT / "data"
OUTPUT = DATA / "ipeds_residence_2024.json"
DATA_URL = "https://nces.ed.gov/ipeds/data-generator?year=2024&tableName=EF2024C&HasRV=0&type=csv"
DICTIONARY_URL = "https://nces.ed.gov/ipeds/dictionary-generator?year=2024&tableName=EF2024C"
CACHE = Path(tempfile.gettempdir()) / "enrollment-squeeze-ipeds"

STATE_FIPS = {
    "AL": "1", "AK": "2", "AZ": "4", "AR": "5", "CA": "6", "CO": "8",
    "CT": "9", "DE": "10", "DC": "11", "FL": "12", "GA": "13", "HI": "15",
    "ID": "16", "IL": "17", "IN": "18", "IA": "19", "KS": "20", "KY": "21",
    "LA": "22", "ME": "23", "MD": "24", "MA": "25", "MI": "26", "MN": "27",
    "MS": "28", "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33",
    "NJ": "34", "NM": "35", "NY": "36", "NC": "37", "ND": "38", "OH": "39",
    "OK": "40", "OR": "41", "PA": "42", "RI": "44", "SC": "45", "SD": "46",
    "TN": "47", "TX": "48", "UT": "49", "VT": "50", "VA": "51", "WA": "53",
    "WV": "54", "WI": "55", "WY": "56",
}
FIPS_STATE = {code: abbreviation for abbreviation, code in STATE_FIPS.items()}
FOREIGN_CODES = {"90"}
UNKNOWN_CODES = {"57", "98"}
OUTLYING_CODES = {"60", "64", "66", "68", "69", "70", "72", "78"}
AGGREGATE_CODES = {"58", "89", "99"}


def classify_residence(code: str, label: str) -> str | None:
    """Classify an EF-C code after validating its official dictionary label."""
    code = str(code)
    expected_special = {
        "57": "State unknown",
        "58": "US total",
        "89": "Outlying areas total",
        "90": "Foreign countries",
        "98": "Residence not reported (balance line)",
        "99": "All first-time degree/certificate-seeking undergraduates total",
    }
    if code in expected_special and label != expected_special[code]:
        raise ValueError(f"Unexpected EF2024C label for code {code}: {label}")
    if code in AGGREGATE_CODES:
        return None
    if code in FIPS_STATE:
        return FIPS_STATE[code]
    if code in FOREIGN_CODES:
        return "FOREIGN"
    if code in UNKNOWN_CODES:
        return "UNKNOWN"
    if code in OUTLYING_CODES:
        return "OUTLYING"
    raise ValueError(f"Unrecognized EF2024C residence code: {code} ({label})")


def aggregate_residence_rows(rows: list[dict], institution_states: dict[str, str]) -> list[dict]:
    """Aggregate standardized residence rows into one record per institution."""
    totals: dict[str, dict[str, int]] = defaultdict(
        lambda: {"home": 0, "other": 0, "foreign": 0, "unknown": 0, "outlying": 0}
    )
    for row in rows:
        unitid = str(row["unitid"])
        count = int(row["count"])
        if count < 0:
            raise ValueError("Residence counts must be nonnegative")
        residence = str(row["residence"])
        if residence == "FOREIGN":
            totals[unitid]["foreign"] += count
        elif residence == "UNKNOWN":
            totals[unitid]["unknown"] += count
        elif residence == "OUTLYING":
            totals[unitid]["outlying"] += count
        elif residence == institution_states.get(unitid):
            totals[unitid]["home"] += count
        elif residence in STATE_FIPS:
            totals[unitid]["other"] += count

    output = []
    for unitid, values in sorted(totals.items()):
        known_domestic = values["home"] + values["other"]
        output.append(
            {
                "unitid": unitid,
                "firstTimeHomeStateCount": values["home"],
                "firstTimeOtherStateCount": values["other"],
                "firstTimeForeignCountryCount": values["foreign"],
                "firstTimeUnknownResidenceCount": values["unknown"],
                "firstTimeOtherUSAreaCount": values["outlying"],
                "firstTimeKnownDomesticCount": known_domestic,
                "firstTimeHomeStateShare": values["home"] / known_domestic if known_domestic else None,
                "firstTimeOtherStateShare": values["other"] / known_domestic if known_domestic else None,
                "residenceSourceYear": 2024,
            }
        )
    return output


def download(url: str, destination: Path) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if not destination.exists():
        request = urllib.request.Request(url, headers={"User-Agent": "EnrollmentSqueeze/1.0"})
        with urllib.request.urlopen(request, timeout=90) as response:
            destination.write_bytes(response.read())
    return destination


def csv_member(zip_path: Path) -> tuple[str, bytes]:
    with zipfile.ZipFile(zip_path) as archive:
        members = [name for name in archive.namelist() if name.lower().endswith(".csv")]
        if len(members) != 1:
            raise ValueError(f"Expected one EF2024C CSV, found {members}")
        return members[0], archive.read(members[0])


def dictionary_labels(zip_path: Path) -> dict[str, str]:
    """Read EFCSTATE labels from the official dictionary workbook ZIP."""
    from openpyxl import load_workbook

    with zipfile.ZipFile(zip_path) as archive:
        members = [name for name in archive.namelist() if name.lower().endswith(".xlsx")]
        if len(members) != 1:
            raise ValueError(f"Expected one EF2024C dictionary workbook, found {members}")
        workbook = load_workbook(BytesIO(archive.read(members[0])), read_only=True, data_only=True)
    labels = {
        str(row[3]): str(row[4])
        for row in workbook["Frequencies"].iter_rows(values_only=True)
        if row[0] == "EFCSTATE"
    }
    if not {"1", "57", "58", "90", "98", "99"}.issubset(labels):
        raise ValueError("EF2024C dictionary is missing expected EFCSTATE labels")
    return labels


def source_rows(zip_path: Path, labels: dict[str, str]) -> list[dict]:
    _, raw = csv_member(zip_path)
    reader = csv.DictReader(io.StringIO(raw.decode("utf-8-sig")))
    required = {"UNITID", "EFCSTATE", "EFRES01"}
    if not required.issubset(reader.fieldnames or []):
        raise ValueError(f"EF2024C is missing required fields: {required}")
    rows = []
    for row in reader:
        code = str(row["EFCSTATE"]).strip()
        residence = classify_residence(code, labels.get(code, ""))
        if residence is None:
            continue
        raw_count = str(row["EFRES01"] or "").replace(",", "").strip()
        count = int(float(raw_count)) if raw_count else 0
        rows.append({"unitid": str(row["UNITID"]), "residence": residence, "count": count})
    return rows


def institution_states() -> dict[str, str]:
    institutions = json.loads((DATA / "institutions.json").read_text(encoding="utf-8"))
    states: dict[str, str] = {}
    for row in sorted(institutions, key=lambda item: item.get("year", 0), reverse=True):
        unitid = str(row["unitid"])
        if unitid not in states and row.get("state") in STATE_FIPS:
            states[unitid] = row["state"]
    return states


def build() -> dict:
    data_zip = download(DATA_URL, CACHE / "EF2024C.zip")
    dictionary_zip = download(DICTIONARY_URL, CACHE / "EF2024C_dictionary.zip")
    states = institution_states()
    labels = dictionary_labels(dictionary_zip)
    records = [row for row in aggregate_residence_rows(source_rows(data_zip, labels), states) if row["unitid"] in states]
    payload = {
        "source": "NCES IPEDS",
        "table": "EF2024C",
        "collection": "Fall 2024",
        "release": "Provisional",
        "population": "First-time degree/certificate-seeking undergraduates",
        "denominator": "Students with known U.S. state or District of Columbia residence",
        "sourceUrl": DATA_URL,
        "records": records,
    }
    OUTPUT.write_text(json.dumps(payload, separators=(",", ":"), allow_nan=False), encoding="utf-8")
    print(f"Built {len(records):,} Fall 2024 institution residence records")
    return payload


if __name__ == "__main__":
    build()
